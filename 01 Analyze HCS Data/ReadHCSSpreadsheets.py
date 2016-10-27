# -*- coding: utf-8 -*-
"""
Created on Thu Oct 27 10:34:55 2016

This script reads HCS spreadsheet data. Before running this script, aggregate
relevant HCS spreadsheets and convert to an Excel spreadsheet.

The script reads the following parameters from the HCS spreadsheets:
- Volume
- Heavy duty percent
- Bus volumes
- Flow rate adjusted volumes
- Control delay

To run this script on path data, use the following command:
python ReadHCSSpreadsheets.py Build_PM.xlsx _Build_SAT -i "Raw Data" -o "Organized Data"

@author: thasegawa
"""

import os
import argparse
import numpy as np
import openpyxl

# Function to check if row difference value is correct, based on criteria defined by text
def checkrowdiff(ws, firstrow_list, rowdiff, text):
    for firstrow in firstrow_list:
        if ws.cell(row = firstrow + rowdiff + 1, column = 1).value != text:
            print('Row difference does not meet criteria ({0}) for firstrow: {1}'.format(text, firstrow))
    
# Function to read in data using firstrow_list, rowdiff, and col_list
def readdata(ws, firstrow_list, col_list, rowdiff, adj_list = False):
    out = np.zeros((len(firstrow_list), len(col_list)))
    
    # Iterate through rows and columns
    for rowindex, firstrow in enumerate(firstrow_list):
        for colindex, col in enumerate(col_list):           
            # Retrieve cell value
            rownum = firstrow + rowdiff + 1
            colnum = openpyxl.cell.column_index_from_string(col)
            val = ws.cell(row = rownum, column = colnum).value
            
            # If cell is empty look in neighboring cells
            if (val is None) and adj_list:
                for adj in adj_list:
                    try:
                        val = ws.cell(row = rownum, column = colnum + adj).value
                    except:
                        continue
                    if val is not None:
                        break
                    
            # If cell value is not empty, add to output
            if val is not None:
                out[rowindex][colindex] = val
    
    return out
    
# Function to write numpy matrix data
def writedata(table, outfname, outdir = '', colheaders = False):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add headers if it was specified in input    
    if colheaders:
        ws.append(colheaders)
        
    # Add data to worksheet, row by row
    tableshape = np.shape(table)
    for i in range(tableshape[0]):
        ws.append(list(table[i]))
    
    # Output worksheet
    outpath = os.path.join(outdir, outfname)
    wb.save(filename = outpath)
            
# =============================================================================

print('Running ReadHCSSpreadsheets.py...')

# Parse arguments
parser = argparse.ArgumentParser()
parser.add_argument('infile')
parser.add_argument('outfname_suffix')
parser.add_argument('-i', '--indir', default='', type=str, help="input directory")
parser.add_argument('-o', '--outdir', default='', type=str, help="output directory")
args = vars(parser.parse_args())
infile, outfname_suffix, indir, outdir = args['infile'], args['outfname_suffix'], args['indir'], args['outdir']

print('Reading data from {0}'.format(infile))
print('Output suffix: {0}'.format(outfname_suffix))
if outdir:
    print('Output directory: {0}'.format(outdir))

# Read Excel file
wb = openpyxl.load_workbook(os.path.join(indir, infile))
ws = wb.worksheets[0]

# Find first row in each HCS spreadsheet
firstrow_list = [rownum for rownum in range(ws.max_row) if ws.cell(row = rownum + 1, column = 1).value == 'LONG REPORT']
numint = len(firstrow_list)

# Define row differences (from first row) for various values
rowdiff_vol = 8
rowdiff_hdperc = 9
rowdiff_busvol = 20
rowdiff_adjvol = 61
rowdiff_delay = 83

# Double check if row differences are correct
checkrowdiff(ws, firstrow_list, rowdiff_vol, 'Volume (vph)')
checkrowdiff(ws, firstrow_list, rowdiff_hdperc, '% Heavy Veh')
checkrowdiff(ws, firstrow_list, rowdiff_busvol, 'Bus Stops/Hour')
checkrowdiff(ws, firstrow_list, rowdiff_adjvol, 'Adjusted Flow Rate')
checkrowdiff(ws, firstrow_list, rowdiff_delay, 'Control Delay')

# Specify column headers
colheaders = ['{0} - {1}'.format(direction, turn) for direction in ['EB', 'WB', 'NB', 'SB'] for turn in ['LT', 'TH', 'RT']]

# Specify column numbers for volume, heavy duty percent, and bus volume
col_list = ['H', 'L', 'R', 'V', 'AB', 'AG', 'AK', 'AQ', 'AT', 'AX', 'BA', 'BD']

# Read and write volume data
print('Reading and writing volume data...')
vol = readdata(ws, firstrow_list, col_list, rowdiff_vol)
writedata(vol, 'vol{0}.xlsx'.format(outfname_suffix), outdir = outdir, colheaders = colheaders)

# Read and write heavy duty percent data
print('Reading and writing heavy duty percent data...')
hdperc = readdata(ws, firstrow_list, col_list, rowdiff_hdperc)
writedata(hdperc, 'hdperc{0}.xlsx'.format(outfname_suffix), outdir = outdir, colheaders = colheaders)

# Read and write bus volume data
print('Reading and writing bus volume data...')
busvol = readdata(ws, firstrow_list, col_list, rowdiff_busvol)
writedata(busvol, 'busvol{0}.xlsx'.format(outfname_suffix), outdir = outdir, colheaders = colheaders)

# Specify column numbers for flow rate adjusted volumes and control delay
col_list = ['E', 'I', 'O', 'S', 'Y', 'AE', 'AI', 'AP', 'AS', 'AX', 'BA', 'BD']

# Read and write flow rate adjusted volumes
print('Reading and writing flow rate adjusted volume data...')
adjvol = readdata(ws, firstrow_list, col_list, rowdiff_adjvol)
writedata(adjvol, 'adjvol{0}.xlsx'.format(outfname_suffix), outdir = outdir, colheaders = colheaders)

# Read and write control delay
print('Reading and writing control delay data...')
delay = readdata(ws, firstrow_list, col_list, rowdiff_delay)
writedata(delay, 'delay{0}.xlsx'.format(outfname_suffix), outdir = outdir, colheaders = colheaders)

print('Finished!!!\n')